
#bibliotecas importadas
import tkinter as tk
import customtkinter as ctk
import openpyxl, xlrd
from tkinter import messagebox
from tkinter import StringVar
from tkinter import END
import pathlib
from openpyxl import Workbook
# talvez enquanto eu não use a biblioteca, não fica destacado, vamos ver como será
# vamos começar o app, vou descrever todos os passos

#criando as aparências do sistema
ctk.set_appearance_mode("system")
ctk.set_default_color_theme("blue")#padrão azul da janela





class App(ctk.CTk):
    def __init__(self):
        super().__init__()#super é a classe principal de sistema
        self.layout_config()
        self.aparencia()
        self.todo_sistema()

    def layout_config(self):
        self.title("Planner de cadastro de jogos")
        self.geometry("700x500")

    def aparencia(self):
       self.lb_apm = ctk.CTkLabel(self, text="Tema", bg_color= "transparent", text_color=['#000','#fff']).place(x=30,y=430) #modo escuro e light.
       self.opt_apm = ctk.CTkOptionMenu(self, values=['Normal','dark'],command=self.change_apm).place(x=30,y=460)

    def todo_sistema(self):
        frame = ctk.CTkFrame(self, width=700,height=50, corner_radius=0, bg_color='teal',fg_color='teal').place(x=0,y=10)
        title = ctk.CTkLabel(frame, text="Planner de Games", font=('Century Gothic bold',28),fg_color='teal',text_color='#fff').place(x=235,y=20)
        spam = ctk.CTkLabel(self,text='Finalizou mais um jogo? Preencha o cadastro para manter suas metas organizadas', font=('Century Gothic bold',16), text_color=['#000','#fff']).place(x=30,y=70)


        ficheiro = pathlib.Path('jogos.xlsx')  # criei a tabela
        if ficheiro.exists():
            pass
        else:
            ficheiro=Workbook()
            planilha = ficheiro.active
            planilha['A1'] = 'Nome do Jogo'
            planilha['B1'] = 'Ano de lançamento'
            planilha['C1'] = 'Tempo'
            planilha['D1'] = 'Plataforma'
            planilha['E1'] = 'Estilo'
            planilha['F1'] = 'Dificuldade'
            planilha['G1'] = 'Nota'
            planilha['H1'] = 'Observações'

            ficheiro.save('jogos.xlsx')


    #criando as funções de salvar e limpar dentro da função
        def salvar(): #pegar os dados dos entrys


            nome= nome_value.get()
            ano = ano_value.get()
            tempo = tempo_value.get()
            plataforma = plataforma_combobox.get()
            estilo = estilo_combobox.get()
            dificuldade = dificuldade_combobox.get()
            nota = nota_combobox.get()
            obs = obs_entry.get(0.0, END)

            if (nome =='' or ano=='' or tempo==''):
                messagebox.showerror('Sistema','ERRO!\nPor favor,prencha os campos!')
            else:

                #Adicionando os dados na tabela via OPENPYXL
                ficheiro = openpyxl.load_workbook('jogos.xlsx')
                planilha=ficheiro.active
                planilha.cell(column=1,row=planilha.max_row+1,value=nome) #sempre cria +1 linha no restante não precisa colocar
                planilha.cell(column=2,row=planilha.max_row, value=ano)
                planilha.cell(column=3, row=planilha.max_row, value=tempo)
                planilha.cell(column=4, row=planilha.max_row, value=plataforma)
                planilha.cell(column=5, row=planilha.max_row, value=estilo)
                planilha.cell(column=6, row=planilha.max_row, value=dificuldade)
                planilha.cell(column=7, row=planilha.max_row, value=nota)
                planilha.cell(column=8, row=planilha.max_row, value=obs)
                ficheiro.save(r'jogos.xlsx')
                messagebox.showinfo('Sistema','Dados salvos com sucesso!')

        def clear():
            nome_value.set("")
            ano_value.set("")
            tempo_value.set("")
            plataforma_combobox.set("")
            estilo_combobox.set("")
            dificuldade_combobox.set("")
            nota_combobox.set("")
            obs_entry.delete(0.0, END)


        #variaveis de texto
        nome_value = StringVar()
        ano_value= StringVar()
        tempo_value= StringVar()


        #entrys
        nome_entry=ctk.CTkEntry(self,width=310,textvariable=nome_value,font=('Century Gothic',16),fg_color='transparent')
        ano_entry = ctk.CTkEntry(self, width=100,textvariable=ano_value, font=('Century Gothic', 16), fg_color='transparent')
        tempo_entry = ctk.CTkEntry(self, width=100,textvariable=tempo_value, font=('Century Gothic', 16), fg_color='transparent')

        #comboboxs
        plataforma_combobox = ctk.CTkComboBox(self,values=['Plastation4','Switch','SuperNintendo','Megadrive','Nintendinho','Nintendo64','Gameboys','PC','Outros'],font=('Century Gothic bold',14))
        status_combobox = ctk.CTkComboBox(self, values=['Finalizado','Abandonado','Estou travado','Em andamento'],font=('Century Gothic bold', 14))
        estilo_combobox = ctk.CTkComboBox(self, values=['Ação', 'Plataforma', 'Rpg', 'Esporte','Luta','Corrida','FPS/Shooter','Cardgame','Simulador','Puzzle','Storytelling','Outros'],font=('Century Gothic bold', 14))
        dificuldade_combobox = ctk.CTkComboBox(self, values=['S+', 'S', 'A', 'B','C'],font=('Century Gothic bold', 14))
        nota_combobox = ctk.CTkComboBox(self, values=['1-Tragédia', '2-Terrível', '3-Ruim', '4-Medíocre','5-Tanto faz','6-Decente','7-Bom','8-Muito bom','9-Ótimo','10-Incrível','11-Jogo da vida'],font=('Century Gothic bold', 14))
        #entrada de textos - observações
        obs_entry = ctk.CTkTextbox(self,width=290,height=228,font=('arial',18),border_color='#aaa',border_width=2,fg_color='transparent' )

        #labels
        nome_lb = ctk.CTkLabel(self,text='Nome do jogo:',font=('Century Gothic bold',16), text_color=['#000','#fff'])
        plataforma_lb = ctk.CTkLabel(self, text='Plataforma:', font=('Century Gothic bold', 16),text_color=['#000', '#fff'])
        status_lb = ctk.CTkLabel(self, text='Status:', font=('Century Gothic bold', 16), text_color=['#000', '#fff'])
        ano_lb = ctk.CTkLabel(self, text='Ano de lançamento:', font=('Century Gothic bold', 16), text_color=['#000', '#fff'])
        estilo_lb = ctk.CTkLabel(self, text='Estilo:', font=('Century Gothic bold', 16), text_color=['#000', '#fff'])
        tempo_lb = ctk.CTkLabel(self, text='Tempo:', font=('Century Gothic bold', 16), text_color=['#000', '#fff'])
        dificuldade_lb = ctk.CTkLabel(self, text='Dificuldade:', font=('Century Gothic bold', 16), text_color=['#000', '#fff'])
        nota_lb = ctk.CTkLabel(self, text='Nota:', font=('Century Gothic bold', 16), text_color=['#000', '#fff'])
        observação_lb = ctk.CTkLabel(self, text='Observações:', font=('Century Gothic bold', 16), text_color=['#000', '#fff'])

        #criação dos botões
        botao_salvar = ctk.CTkButton(self,text="Salvar".upper(),command=salvar,fg_color="#151",hover_color='#131').place(x= 350,y=400)   #função hover_color é para mudar de cor quando a seta ficar em cima do botão
        botao_limpar = ctk.CTkButton(self, text="Limpar".upper(),command=clear, fg_color="#155", hover_color='#333').place(x=500, y= 400)

        #posicionamento dos elementos na tela
        nome_lb.place(x=30,y=100)
        nome_entry.place(x=30,y=130)
        plataforma_lb.place(x=30,y=160)
        plataforma_combobox.place(x=30,y=190)
        estilo_lb.place(x=200,y=160)
        estilo_combobox.place(x=200, y=190)
        dificuldade_lb.place(x=30,y=225)
        dificuldade_combobox.place(x=30,y=260)
        nota_lb.place(x=200, y=225)
        nota_combobox.place(x=200, y=260)
        ano_lb.place(x=30,y=295)
        ano_entry.place(x=30,y=330)
        tempo_lb.place(x=200 ,y=295)
        tempo_entry.place(x=200,y=330)
        observação_lb.place(x=350,y=100)
        obs_entry.place(x=350,y=130)


    def change_apm(self, nova_aparencia):
        ctk.set_appearance_mode(nova_aparencia)
if __name__=="__main__":
    app = App()
    app.mainloop()

